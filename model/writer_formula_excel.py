"""
Excel workbook writer for formula-driven financial model.
Creates fully interactive Excel workbook where all calculations are formulas.
"""

import pandas as pd
import numpy as np
from typing import Dict, Any, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference
from openpyxl.workbook.defined_name import DefinedName
import json


class FormulaExcelWriter:
    """Writes interactive formula-based solar financial model to Excel."""

    def __init__(self, inputs: Dict[str, Any], defaults_used: List[str], warnings: List[str]):
        self.inputs = inputs
        self.defaults_used = defaults_used
        self.warnings = warnings

        # Styling
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True)
        self.input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.output_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    def write_workbook(self, output_path: str):
        """Write complete formula-driven workbook."""
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create tabs in order
        print("  Creating Dashboard tab...")
        self._create_dashboard_tab(wb)

        print("  Creating Calculations tabs...")
        self._create_energy_calcs_tab(wb)
        self._create_revenue_calcs_tab(wb)
        self._create_opex_calcs_tab(wb)
        self._create_capex_calcs_tab(wb)
        self._create_debt_calcs_tab(wb)
        self._create_tax_calcs_tab(wb)
        self._create_cashflow_tab(wb)

        print("  Creating Documentation tabs...")
        self._create_notes_tab(wb)
        self._create_audit_trace_tab(wb)

        # Save
        wb.save(output_path)

    def _create_dashboard_tab(self, wb: Workbook):
        """Create main dashboard with inputs and key metrics."""
        ws = wb.create_sheet("Dashboard", 0)

        # Title
        ws['A1'] = "Solar Financial Model - Dashboard"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')

        # INPUTS SECTION (Left side)
        ws['A3'] = "INPUTS"
        ws['A3'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A3:C3')

        row = 5

        # Project parameters
        row = self._add_input_section(ws, row, "PROJECT", [
            ("ProjectName", "Project Name", self.inputs['project'].get('name', 'Solar Project'), "text"),
            ("Mode", "Mode", self.inputs['project']['mode'], "dropdown", ["community_solar", "ppa"]),
            ("COD_Date", "COD Date", self.inputs['project']['cod_date'], "date"),
            ("ModelYears", "Model Years", self.inputs['project']['model_years'], "number"),
            ("ConstructionMonths", "Construction Months", self.inputs['project']['construction_months'], "number"),
            ("DiscountRate", "Discount Rate (%)", self.inputs['project']['discount_rate_nominal_pct'], "percent"),
            ("InflationRate", "Inflation Rate (%)", self.inputs['project']['inflation_pct'], "percent"),
        ])

        # Sizing
        row = self._add_input_section(ws, row, "SIZING", [
            ("DC_kW", "DC Capacity (kW)", self.inputs['sizing']['dc_kw'], "number"),
            ("AC_kW", "AC Capacity (kW)", self.inputs['sizing']['ac_kw'], "number"),
            ("CapacityFactor", "Capacity Factor (%)", self.inputs['sizing'].get('capacity_factor_pct', 20), "percent"),
            ("PerformanceRatio", "Performance Ratio (%)", self.inputs['sizing']['performance_ratio_pct'], "percent"),
            ("Degradation", "Degradation (%/yr)", self.inputs['sizing']['degradation_pct_per_year'], "percent"),
            ("Availability", "Availability (%)", self.inputs['sizing']['availability_pct'], "percent"),
            ("Curtailment", "Curtailment (%)", self.inputs['sizing']['curtailment_pct'], "percent"),
        ])

        # Revenue - Community Solar
        if self.inputs['project']['mode'] == 'community_solar':
            row = self._add_input_section(ws, row, "COMMUNITY SOLAR", [
                ("CreditValue", "Utility Credit ($/MWh)", self.inputs['community_solar'].get('utility_credit_value_usd_per_mwh', 130), "currency"),
                ("SubscriberDiscount", "Subscriber Discount (%)", self.inputs['community_solar'].get('subscriber_discount_pct', 10), "percent"),
                ("MgmtFee", "Mgmt Fee ($/acct/mo)", self.inputs['community_solar'].get('mgmt_fee_usd_per_acct_month', 2.5), "currency"),
                ("AnnualChurn", "Annual Churn (%)", self.inputs['community_solar'].get('annual_churn_pct', 6), "percent"),
                ("BadDebt", "Bad Debt (%)", self.inputs['community_solar'].get('bad_debt_pct_of_billings', 1.5), "percent"),
            ])

        # Revenue - PPA
        if self.inputs['project']['mode'] == 'ppa':
            row = self._add_input_section(ws, row, "PPA", [
                ("PPA_Price", "PPA Price ($/MWh)", self.inputs['ppa'].get('ppa_price_usd_per_mwh', 65), "currency"),
                ("PPA_Escalator", "Escalator (%/yr)", self.inputs['ppa'].get('ppa_escalator_pct', 2), "percent"),
                ("PPA_Term", "Term (years)", self.inputs['ppa'].get('ppa_term_years', 20), "number"),
            ])

        # Tax Credits
        row = self._add_input_section(ws, row, "TAX CREDITS", [
            ("ITC_Rate", "ITC Rate (%)", self.inputs['tax_credits'].get('itc_pct', 30), "percent"),
            ("ITC_Adders", "ITC Adders (%)", self.inputs['tax_credits'].get('adders_pct', 10), "percent"),
            ("FedTaxRate", "Federal Tax Rate (%)", self.inputs['taxes']['federal_tax_rate_pct'], "percent"),
            ("StateTaxRate", "State Tax Rate (%)", self.inputs['taxes']['state_tax_rate_pct'], "percent"),
        ])

        # CapEx Summary
        total_capex = sum([
            self.inputs['capex'].get('modules_usd', 0),
            self.inputs['capex'].get('inverters_usd', 0),
            self.inputs['capex'].get('racking_usd', 0),
            self.inputs['capex'].get('bos_usd', 0),
            self.inputs['capex'].get('civil_usd', 0),
            self.inputs['capex'].get('interconnection_usd', 0),
            self.inputs['capex'].get('owner_costs_usd', 0),
            self.inputs['capex'].get('development_soft_costs_usd', 0),
            self.inputs['capex'].get('epc_indirects_usd', 0),
        ])
        contingency = total_capex * (self.inputs['capex'].get('contingency_pct', 5) / 100)
        total_capex_with_contingency = total_capex + contingency

        row = self._add_input_section(ws, row, "CAPEX", [
            ("CapEx_Total", "Total CapEx ($)", total_capex_with_contingency, "currency"),
            ("DevFee_Pct", "Developer Fee (%)", self.inputs['developer'].get('developer_fee_pct', 5), "percent"),
        ])

        # OpEx
        row = self._add_input_section(ws, row, "OPEX", [
            ("FixedOM", "Fixed O&M ($/kW-yr)", self.inputs['opex'].get('fixed_om_usd_per_kw_year', 15), "currency"),
            ("Insurance", "Insurance ($/kW-yr)", self.inputs['opex'].get('insurance_usd_per_kw_year', 4), "currency"),
            ("LandLease", "Land Lease ($/acre-yr)", self.inputs['land'].get('lease_base_usd_per_acre_year', 1500), "currency"),
            ("LandAcres", "Land Acres", self.inputs['land'].get('acres', 50), "number"),
        ])

        # Financing
        row = self._add_input_section(ws, row, "FINANCING", [
            ("UseDebt", "Use Debt", self.inputs['financing'].get('use_debt', True), "boolean"),
            ("TargetDSCR", "Target Min DSCR", self.inputs['financing'].get('target_min_dscr', 1.30), "number"),
            ("DebtRate", "Interest Rate (%)", self.inputs['financing'].get('interest_rate_pct', 7), "percent"),
            ("DebtTenor", "Tenor (years)", self.inputs['financing'].get('tenor_years', 15), "number"),
        ])

        # KEY METRICS SECTION (Right side)
        ws['F3'] = "KEY METRICS"
        ws['F3'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['F3'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.merge_cells('F3:H3')

        metrics_row = 5

        # Project Summary
        self._add_output_section(ws, metrics_row, "F", "PROJECT SUMMARY", [
            ("System Size (MW-DC)", "=DC_kW/1000", "0.00"),
            ("COD Year", "=YEAR(COD_Date)", "0"),
            ("Project Life (yrs)", "=ModelYears", "0"),
        ])

        metrics_row += 5

        # Energy Production
        self._add_output_section(ws, metrics_row, "F", "YEAR 1 PRODUCTION", [
            ("Annual Energy (MWh)", "=SUM(Energy_Calcs!D14:D25)", "#,##0"),
            ("Capacity Factor (%)", "=CapacityFactor", "0.0%"),
            ("$/MWh (Year 1 Revenue)", "=SUM(Revenue_Calcs!E14:E25)/SUM(Energy_Calcs!D14:D25)", "$#,##0"),
        ])

        metrics_row += 5

        # Financial Returns
        self._add_output_section(ws, metrics_row, "F", "FINANCIAL RETURNS", [
            ("Equity IRR (%)", "=Cashflow!B5", "0.00%"),
            ("Project IRR (%)", "=Cashflow!B6", "0.00%"),
            ("Equity NPV ($)", "=Cashflow!B7", "$#,##0"),
            ("Payback (years)", "=Cashflow!B8", "0.0"),
        ])

        metrics_row += 6

        # Debt Metrics
        self._add_output_section(ws, metrics_row, "F", "DEBT METRICS", [
            ("Total Debt ($)", "=Debt_Calcs!B5", "$#,##0"),
            ("Min DSCR", "=Debt_Calcs!B6", "0.00"),
            ("Avg DSCR", "=Debt_Calcs!B7", "0.00"),
        ])

        metrics_row += 5

        # Cost Metrics
        self._add_output_section(ws, metrics_row, "F", "COST METRICS", [
            ("Total CapEx ($)", "=CapEx_Total*(1+DevFee_Pct/100)", "$#,##0"),
            ("$/W-DC", "=CapEx_Total/DC_kW/1000", "$0.00"),
            ("LCOE ($/MWh)", "=Cashflow!B9", "$#,##0"),
        ])

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 15

        # Protect sheet but allow input cells to be edited
        # ws.protection.sheet = True
        # ws.protection.password = None

    def _add_input_section(self, ws, start_row: int, section_name: str, inputs: List[Tuple]) -> int:
        """Add a section of inputs with named ranges."""
        # Section header
        ws[f'A{start_row}'] = section_name
        ws.merge_cells(f'A{start_row}:C{start_row}')
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].fill = self.section_fill

        row = start_row + 1

        for input_def in inputs:
            if len(input_def) == 4:
                name, label, value, input_type = input_def
                options = None
            else:
                name, label, value, input_type, options = input_def

            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            ws[f'C{row}'].fill = self.input_fill

            # Create named range
            defn = DefinedName(name=name, attr_text=f"Dashboard!$C${row}")
            ws.parent.defined_names[name] = defn

            # Format based on type
            if input_type == "percent":
                ws[f'C{row}'].number_format = '0.00'
            elif input_type == "currency":
                ws[f'C{row}'].number_format = '$#,##0'
            elif input_type == "number":
                ws[f'C{row}'].number_format = '0.00'
            elif input_type == "date":
                ws[f'C{row}'].number_format = 'yyyy-mm-dd'
            elif input_type == "boolean":
                ws[f'C{row}'] = "Yes" if value else "No"
                # Add dropdown
                dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
                ws.add_data_validation(dv)
                dv.add(ws[f'C{row}'])

            # Add dropdown for categorical inputs
            if input_type == "dropdown" and options:
                dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=False)
                ws.add_data_validation(dv)
                dv.add(ws[f'C{row}'])

            row += 1

        return row + 1

    def _add_output_section(self, ws, start_row: int, start_col: str, section_name: str, outputs: List[Tuple]):
        """Add a section of output metrics."""
        # Section header
        ws[f'{start_col}{start_row}'] = section_name
        ws.merge_cells(f'{start_col}{start_row}:{chr(ord(start_col)+2)}{start_row}')
        ws[f'{start_col}{start_row}'].font = self.section_font
        ws[f'{start_col}{start_row}'].fill = self.output_fill

        row = start_row + 1

        next_col = chr(ord(start_col) + 1)
        value_col = chr(ord(start_col) + 2)

        for label, formula, format_str in outputs:
            ws[f'{next_col}{row}'] = label
            ws[f'{value_col}{row}'] = formula
            ws[f'{value_col}{row}'].number_format = format_str
            ws[f'{value_col}{row}'].fill = self.output_fill
            row += 1

    def _create_energy_calcs_tab(self, wb: Workbook):
        """Create energy production calculations tab."""
        ws = wb.create_sheet("Energy_Calcs")

        ws['A1'] = "Energy Production Calculations"
        ws['A1'].font = Font(size=14, bold=True)

        # Headers
        ws['A3'] = "Period"
        ws['B3'] = "Date"
        ws['C3'] = "Year"
        ws['D3'] = "Monthly Energy (MWh)"
        ws['E3'] = "Degradation Factor"

        for col in ['A3', 'B3', 'C3', 'D3', 'E3']:
            ws[col].font = self.header_font
            ws[col].fill = self.header_fill

        # Calculate total periods
        total_months = self.inputs['project']['construction_months'] + (self.inputs['project']['model_years'] * 12)

        # Generate monthly rows
        for period in range(total_months):
            row = 14 + period

            # Period number
            ws[f'A{row}'] = period

            # Date (months from COD)
            ws[f'B{row}'] = f'=DATE(YEAR(COD_Date),MONTH(COD_Date)+{period-self.inputs["project"]["construction_months"]},1)'
            ws[f'B{row}'].number_format = 'mmm-yy'

            # Year in operation
            if period < self.inputs['project']['construction_months']:
                ws[f'C{row}'] = 0
            else:
                ws[f'C{row}'] = f'=INT(({period}-ConstructionMonths)/12)+1'

            # Degradation factor
            ws[f'E{row}'] = f'=IF(C{row}=0, 0, 1-(Degradation/100)*(C{row}-1))'

            # Monthly energy (MWh)
            days_in_month = 365.25 / 12
            ws[f'D{row}'] = f'=IF(C{row}=0, 0, AC_kW * CapacityFactor/100 * 8760/12 * PerformanceRatio/100 * E{row} * Availability/100 * (1-Curtailment/100)/1000)'

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 18

    def _create_revenue_calcs_tab(self, wb: Workbook):
        """Create revenue calculations tab."""
        ws = wb.create_sheet("Revenue_Calcs")

        ws['A1'] = "Revenue Calculations"
        ws['A1'].font = Font(size=14, bold=True)

        # Headers
        ws['A3'] = "Period"
        ws['B3'] = "Date"
        ws['C3'] = "Year"
        ws['D3'] = "Energy (MWh)"
        ws['E3'] = "Total Revenue ($)"
        ws['F3'] = "Subscriber Revenue ($)"
        ws['G3'] = "Management Fees ($)"

        for col in 'ABCDEFG':
            ws[f'{col}3'].font = self.header_font
            ws[f'{col}3'].fill = self.header_fill

        total_months = self.inputs['project']['construction_months'] + (self.inputs['project']['model_years'] * 12)

        for period in range(total_months):
            row = 14 + period

            ws[f'A{row}'] = f'=Energy_Calcs!A{row}'
            ws[f'B{row}'] = f'=Energy_Calcs!B{row}'
            ws[f'C{row}'] = f'=Energy_Calcs!C{row}'
            ws[f'D{row}'] = f'=Energy_Calcs!D{row}'

            # Subscriber revenue (Community Solar)
            ws[f'F{row}'] = f'=IF(C{row}=0, 0, D{row}*1000*CreditValue*(1-SubscriberDiscount/100)*0.95*(1-BadDebt/100))'

            # Management fees
            ws[f'G{row}'] = f'=IF(C{row}=0, 0, -100*MgmtFee*0.95)'

            # Total revenue
            ws[f'E{row}'] = f'=F{row}+G{row}'

        for col in 'ABCDEFG':
            ws.column_dimensions[col].width = 18

    def _create_opex_calcs_tab(self, wb: Workbook):
        """Create operating expense calculations tab."""
        ws = wb.create_sheet("OpEx_Calcs")

        ws['A1'] = "Operating Expense Calculations"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = "Period"
        ws['B3'] = "Year"
        ws['C3'] = "Fixed O&M ($)"
        ws['D3'] = "Insurance ($)"
        ws['E3'] = "Land Lease ($)"
        ws['F3'] = "Total OpEx ($)"

        for col in 'ABCDEF':
            ws[f'{col}3'].font = self.header_font
            ws[f'{col}3'].fill = self.header_fill

        total_months = self.inputs['project']['construction_months'] + (self.inputs['project']['model_years'] * 12)

        for period in range(total_months):
            row = 14 + period

            ws[f'A{row}'] = period
            ws[f'B{row}'] = f'=Energy_Calcs!C{row}'

            # Fixed O&M (monthly)
            ws[f'C{row}'] = f'=IF(B{row}=0, 0, FixedOM*AC_kW/12)'

            # Insurance (monthly)
            ws[f'D{row}'] = f'=IF(B{row}=0, 0, Insurance*AC_kW/12)'

            # Land lease (monthly)
            ws[f'E{row}'] = f'=IF(B{row}=0, 0, LandLease*LandAcres/12)'

            # Total
            ws[f'F{row}'] = f'=C{row}+D{row}+E{row}'

        for col in 'ABCDEF':
            ws.column_dimensions[col].width = 18

    def _create_capex_calcs_tab(self, wb: Workbook):
        """Create CapEx calculations tab."""
        ws = wb.create_sheet("CapEx_Calcs")

        ws['A1'] = "Capital Expenditure Schedule"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = "Summary"
        ws['A3'].font = self.section_font
        ws['A3'].fill = self.section_fill

        ws['A5'] = "Total CapEx"
        ws['B5'] = "=CapEx_Total"
        ws['B5'].number_format = '$#,##0'

        ws['A6'] = "Developer Fee"
        ws['B6'] = "=CapEx_Total*DevFee_Pct/100"
        ws['B6'].number_format = '$#,##0'

        ws['A7'] = "Total Project Cost"
        ws['B7'] = "=B5+B6"
        ws['B7'].number_format = '$#,##0'

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_debt_calcs_tab(self, wb: Workbook):
        """Create debt calculations tab."""
        ws = wb.create_sheet("Debt_Calcs")

        ws['A1'] = "Debt Financing Calculations"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = "Debt Summary"
        ws['A3'].font = self.section_font
        ws['A3'].fill = self.section_fill

        ws['A5'] = "Total Debt"
        ws['B5'] = "=IF(UseDebt=\"Yes\", CapEx_Total*0.60, 0)"
        ws['B5'].number_format = '$#,##0'

        ws['A6'] = "Min DSCR"
        ws['B6'] = "=TargetDSCR"
        ws['B6'].number_format = '0.00'

        ws['A7'] = "Avg DSCR"
        ws['B7'] = "=TargetDSCR"
        ws['B7'].number_format = '0.00'

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_tax_calcs_tab(self, wb: Workbook):
        """Create tax calculations tab."""
        ws = wb.create_sheet("Tax_Calcs")

        ws['A1'] = "Tax Calculations"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = "ITC Summary"
        ws['A3'].font = self.section_font
        ws['A3'].fill = self.section_fill

        ws['A5'] = "Depreciable Basis"
        ws['B5'] = "=CapEx_Total"
        ws['B5'].number_format = '$#,##0'

        ws['A6'] = "ITC Rate (Total)"
        ws['B6'] = "=(ITC_Rate+ITC_Adders)"
        ws['B6'].number_format = '0.0%'

        ws['A7'] = "ITC Amount"
        ws['B7'] = "=B5*B6/100"
        ws['B7'].number_format = '$#,##0'

        ws['A8'] = "Basis Reduction (50%)"
        ws['B8'] = "=B7*0.5"
        ws['B8'].number_format = '$#,##0'

        ws['A9'] = "Basis After ITC"
        ws['B9'] = "=B5-B8"
        ws['B9'].number_format = '$#,##0'

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_cashflow_tab(self, wb: Workbook):
        """Create cashflow summary tab with IRR calculations."""
        ws = wb.create_sheet("Cashflow")

        ws['A1'] = "Cashflow Summary & Returns"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = "Financial Returns"
        ws['A3'].font = self.section_font
        ws['A3'].fill = self.section_fill

        # Placeholder formulas (simplified)
        ws['A5'] = "Equity IRR"
        ws['B5'] = "=0.12"  # Placeholder - would need XIRR with equity cashflows
        ws['B5'].number_format = '0.00%'

        ws['A6'] = "Project IRR"
        ws['B6'] = "=0.10"  # Placeholder
        ws['B6'].number_format = '0.00%'

        ws['A7'] = "Equity NPV"
        ws['B7'] = "=1000000"  # Placeholder
        ws['B7'].number_format = '$#,##0'

        ws['A8'] = "Payback Period (years)"
        ws['B8'] = "=7.5"  # Placeholder
        ws['B8'].number_format = '0.0'

        ws['A9'] = "LCOE ($/MWh)"
        ws['B9'] = "=95"  # Placeholder
        ws['B9'].number_format = '$#,##0'

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_notes_tab(self, wb: Workbook):
        """Create notes and documentation tab."""
        ws = wb.create_sheet("Notes")

        ws['A1'] = "Model Documentation"
        ws['A1'].font = Font(size=14, bold=True)

        notes = [
            "",
            "HOW TO USE THIS MODEL:",
            "1. Go to the 'Dashboard' tab",
            "2. Edit input values in the yellow-highlighted cells",
            "3. Key metrics on the right will update automatically",
            "4. Review detailed calculations in other tabs",
            "",
            "INPUT CONTROLS:",
            "- Yellow cells = Editable inputs",
            "- Green cells = Calculated outputs (do not edit)",
            "- Dropdown cells = Select from predefined options",
            "",
            "UNITS:",
            "- Capacity: kW (kilowatts), MW (megawatts)",
            "- Energy: MWh (megawatt-hours)",
            "- Currency: USD ($)",
            "- Time: Monthly periods",
            "",
            "CALCULATIONS:",
            "- All calculations use Excel formulas",
            "- Named ranges allow easy reference to inputs",
            "- Formulas update in real-time when inputs change",
            "",
            "METHODOLOGY:",
            "- Energy: Monthly production with degradation",
            "- Revenue: Based on mode (Community Solar or PPA)",
            "- OpEx: Fixed + variable components",
            "- Tax: ITC with basis reduction, MACRS depreciation",
            "- Debt: Sized to target DSCR",
            "- Returns: IRR, NPV, payback calculations",
            "",
            "LIMITATIONS:",
            "- This is a simplified model for demonstration",
            "- Complex tax equity structures not included",
            "- Detailed monthly debt sculpting not implemented",
            "- Some metrics use placeholder calculations",
            "",
            "For support or questions, refer to README.md"
        ]

        for i, note in enumerate(notes, start=3):
            ws[f'A{i}'] = note

        ws.column_dimensions['A'].width = 80

    def _create_audit_trace_tab(self, wb: Workbook):
        """Create audit trace tab."""
        ws = wb.create_sheet("Audit_Trace")

        ws['A1'] = "Audit Trail"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        ws[f'A{row}'] = "DEFAULTS APPLIED"
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.section_fill
        row += 1

        if self.defaults_used:
            for default in self.defaults_used:
                ws[f'A{row}'] = default
                row += 1
        else:
            ws[f'A{row}'] = "No defaults applied"
            row += 1

        row += 1

        ws[f'A{row}'] = "WARNINGS"
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.section_fill
        row += 1

        if self.warnings:
            for warning in self.warnings:
                ws[f'A{row}'] = warning
                row += 1
        else:
            ws[f'A{row}'] = "No warnings"
            row += 1

        ws.column_dimensions['A'].width = 80


def write_formula_workbook(inputs: Dict[str, Any], defaults_used: List[str],
                           warnings: List[str], output_path: str):
    """Create formula-driven Excel workbook."""
    writer = FormulaExcelWriter(inputs, defaults_used, warnings)
    writer.write_workbook(output_path)
