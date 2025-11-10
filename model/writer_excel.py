"""
Excel workbook writer module.
Generates comprehensive Excel output with all required tabs and formatting.
"""

import pandas as pd
import numpy as np
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
import json


class ExcelWriter:
    """Writes solar financial model to Excel workbook with multiple tabs."""

    def __init__(self, inputs: Dict[str, Any], cashflow: pd.DataFrame,
                 metrics: Dict[str, Any], defaults_used: List[str],
                 warnings: List[str]):
        self.inputs = inputs
        self.cashflow = cashflow
        self.metrics = metrics
        self.defaults_used = defaults_used
        self.warnings = warnings

        # Styling
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True)

    def write_workbook(self, output_path: str):
        """Write complete workbook to file."""
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create all tabs
        self._create_assumptions_tab(wb)
        self._create_inputs_data_tab(wb)
        self._create_energy_tab(wb)
        self._create_revenue_tab(wb)
        self._create_capex_tab(wb)
        self._create_opex_tab(wb)
        self._create_debt_tab(wb)
        self._create_tax_depreciation_tab(wb)
        self._create_cashflow_waterfall_tab(wb)
        self._create_irr_npv_metrics_tab(wb)
        self._create_sensitivities_tab(wb)
        self._create_scenarios_tab(wb)
        self._create_charts_tab(wb)
        self._create_audit_trace_tab(wb)
        self._create_notes_tab(wb)

        # Save
        wb.save(output_path)

    def _create_assumptions_tab(self, wb: Workbook):
        """Create Assumptions tab with high-level summary."""
        ws = wb.create_sheet("Assumptions")

        # Title
        ws['A1'] = "Solar Financial Model - Key Assumptions"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        # Project
        ws[f'A{row}'] = "PROJECT"
        self._apply_section_style(ws[f'A{row}'])
        row += 1

        project = self.inputs['project']
        assumptions = [
            ("Project Name", project.get('name', 'N/A')),
            ("Mode", project['mode']),
            ("COD Date", project['cod_date']),
            ("Model Years", project['model_years']),
            ("Construction Months", project['construction_months']),
            ("Discount Rate", f"{project['discount_rate_nominal_pct']:.2f}%"),
            ("Inflation Rate", f"{project['inflation_pct']:.2f}%"),
        ]

        for label, value in assumptions:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Sizing
        ws[f'A{row}'] = "SIZING"
        self._apply_section_style(ws[f'A{row}'])
        row += 1

        sizing = self.inputs['sizing']
        sizing_data = [
            ("DC Size (kW)", sizing['dc_kw']),
            ("AC Size (kW)", sizing['ac_kw']),
            ("DC/AC Ratio", f"{sizing['dc_ac_ratio']:.2f}"),
            ("Capacity Factor", f"{sizing.get('capacity_factor_pct', 0):.2f}%"),
            ("Performance Ratio", f"{sizing['performance_ratio_pct']:.2f}%"),
            ("Degradation (%/year)", f"{sizing['degradation_pct_per_year']:.3f}%"),
            ("Availability", f"{sizing['availability_pct']:.2f}%"),
            ("Curtailment", f"{sizing['curtailment_pct']:.2f}%"),
        ]

        for label, value in sizing_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Tax Credits
        ws[f'A{row}'] = "TAX CREDITS"
        self._apply_section_style(ws[f'A{row}'])
        row += 1

        tax_credits = self.inputs['tax_credits']
        tax_data = [
            ("Mode", tax_credits['mode']),
            ("ITC %", f"{tax_credits.get('itc_pct', 0):.1f}%"),
            ("Adders %", f"{tax_credits.get('adders_pct', 0):.1f}%"),
            ("Basis Reduction %", f"{tax_credits.get('basis_reduction_pct', 0):.1f}%"),
            ("Bonus Depreciation %", f"{tax_credits.get('bonus_depreciation_pct', 0):.1f}%"),
        ]

        for label, value in tax_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Auto-width columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_inputs_data_tab(self, wb: Workbook):
        """Create Inputs_Data tab with all input parameters in table format."""
        ws = wb.create_sheet("Inputs_Data")

        ws['A1'] = "Input Parameters - Editable Configuration"
        ws['A1'].font = Font(size=14, bold=True)

        # Table headers
        ws['A3'] = "Section"
        ws['B3'] = "Parameter"
        ws['C3'] = "Value"
        ws['D3'] = "Unit / Notes"

        for col in ['A3', 'B3', 'C3', 'D3']:
            self._apply_header_style(ws[col])

        row = 4

        # Helper function to add rows
        def add_section(section_name):
            nonlocal row
            ws[f'A{row}'] = section_name
            ws.merge_cells(f'A{row}:D{row}')
            self._apply_section_style(ws[f'A{row}'])
            row += 1

        def add_param(section, param, value, unit=""):
            nonlocal row
            ws[f'A{row}'] = section
            ws[f'B{row}'] = param
            ws[f'C{row}'] = value
            ws[f'D{row}'] = unit
            row += 1

        # PROJECT
        add_section("PROJECT")
        proj = self.inputs['project']
        add_param("Project", "Name", proj.get('name', 'N/A'), "")
        add_param("Project", "Mode", proj['mode'], "community_solar or ppa")
        add_param("Project", "COD Date", proj['cod_date'], "YYYY-MM-DD")
        add_param("Project", "Model Years", proj['model_years'], "years")
        add_param("Project", "Construction Months", proj['construction_months'], "months")
        add_param("Project", "Discount Rate", proj['discount_rate_nominal_pct'], "% nominal")
        add_param("Project", "Inflation Rate", proj['inflation_pct'], "% per year")
        row += 1

        # SIZING
        add_section("SIZING")
        sizing = self.inputs['sizing']
        add_param("Sizing", "DC Capacity", sizing['dc_kw'], "kW")
        add_param("Sizing", "AC Capacity", sizing['ac_kw'], "kW")
        add_param("Sizing", "DC/AC Ratio", sizing['dc_ac_ratio'], "ratio")
        add_param("Sizing", "Capacity Factor", sizing.get('capacity_factor_pct', 0), "%")
        add_param("Sizing", "Performance Ratio", sizing['performance_ratio_pct'], "%")
        add_param("Sizing", "Degradation Rate", sizing['degradation_pct_per_year'], "% per year")
        add_param("Sizing", "Availability", sizing['availability_pct'], "%")
        add_param("Sizing", "Curtailment", sizing['curtailment_pct'], "%")
        add_param("Sizing", "Use 8760 Data", sizing['use_8760'], "true/false")
        row += 1

        # REVENUE COMMON
        add_section("REVENUE - COMMON")
        rev_common = self.inputs.get('revenue_common', {})
        add_param("Revenue", "REC Price", rev_common.get('rec_price_usd_per_mwh', 0), "$/MWh")
        add_param("Revenue", "REC Escalator", rev_common.get('rec_escalator_pct', 0), "% per year")
        add_param("Revenue", "Capacity Revenue", rev_common.get('capacity_revenue_usd_per_kw_year', 0), "$/kW-year")
        add_param("Revenue", "Capacity Term", rev_common.get('capacity_term_years', 0), "years")
        row += 1

        # COMMUNITY SOLAR
        if self.inputs['project']['mode'] == 'community_solar':
            add_section("COMMUNITY SOLAR")
            cdg = self.inputs.get('community_solar', {})
            add_param("CDG", "Subscriber Discount", cdg.get('subscriber_discount_pct', 0), "%")
            add_param("CDG", "Anchor Share", cdg.get('anchor_share_pct', 0), "%")
            add_param("CDG", "LMI Share", cdg.get('lmi_share_pct', 0), "%")
            add_param("CDG", "Utility Credit Value", cdg.get('utility_credit_value_usd_per_mwh', 0), "$/MWh")
            add_param("CDG", "Management Fee", cdg.get('mgmt_fee_usd_per_acct_month', 0), "$/account/month")
            add_param("CDG", "Acquisition Cost", cdg.get('acquisition_cost_usd_per_subscriber', 0), "$/subscriber")
            add_param("CDG", "Annual Churn", cdg.get('annual_churn_pct', 0), "%")
            add_param("CDG", "Bad Debt", cdg.get('bad_debt_pct_of_billings', 0), "% of billings")
            add_param("CDG", "Ramp to Full Subscribed", cdg.get('ramp_to_full_subscribed_months', 0), "months")
            row += 1

        # PPA
        if self.inputs['project']['mode'] == 'ppa':
            add_section("POWER PURCHASE AGREEMENT")
            ppa = self.inputs.get('ppa', {})
            add_param("PPA", "PPA Price", ppa.get('ppa_price_usd_per_mwh', 0), "$/MWh")
            add_param("PPA", "Escalator", ppa.get('ppa_escalator_pct', 0), "% per year")
            add_param("PPA", "Contract Term", ppa.get('ppa_term_years', 0), "years")
            add_param("PPA", "Settlement", ppa.get('settlement', ''), "pay_as_produced")
            row += 1

        # TAX CREDITS
        add_section("TAX CREDITS")
        tax_cr = self.inputs['tax_credits']
        add_param("Tax", "Mode", tax_cr['mode'], "ITC, PTC, or None")
        add_param("Tax", "ITC Percentage", tax_cr.get('itc_pct', 0), "%")
        add_param("Tax", "Adders", tax_cr.get('adders_pct', 0), "%")
        add_param("Tax", "Basis Reduction", tax_cr.get('basis_reduction_pct', 0), "% of ITC")
        add_param("Tax", "PTC Rate", tax_cr.get('ptc_usd_per_mwh', 0), "$/MWh")
        add_param("Tax", "PTC Term", tax_cr.get('ptc_term_years', 0), "years")
        add_param("Tax", "Bonus Depreciation", tax_cr.get('bonus_depreciation_pct', 0), "%")
        add_param("Tax", "Elective Pay", tax_cr.get('elective_pay', False), "true/false")
        row += 1

        # CAPEX
        add_section("CAPITAL EXPENDITURE")
        capex = self.inputs['capex']
        add_param("CapEx", "Modules", capex.get('modules_usd', 0), "$")
        add_param("CapEx", "Inverters", capex.get('inverters_usd', 0), "$")
        add_param("CapEx", "Racking", capex.get('racking_usd', 0), "$")
        add_param("CapEx", "BOS", capex.get('bos_usd', 0), "$")
        add_param("CapEx", "Civil", capex.get('civil_usd', 0), "$")
        add_param("CapEx", "Interconnection", capex.get('interconnection_usd', 0), "$")
        add_param("CapEx", "Owner Costs", capex.get('owner_costs_usd', 0), "$")
        add_param("CapEx", "Development Soft Costs", capex.get('development_soft_costs_usd', 0), "$")
        add_param("CapEx", "EPC Indirects", capex.get('epc_indirects_usd', 0), "$")
        add_param("CapEx", "Contingency", capex.get('contingency_pct', 0), "% of subtotal")
        add_param("CapEx", "Decommissioning Reserve", capex.get('decommissioning_reserve_usd', 0), "$")
        row += 1

        # DEVELOPER
        add_section("DEVELOPER FEE")
        dev = self.inputs.get('developer', {})
        add_param("Developer", "Fee Mode", dev.get('developer_fee_mode', ''), "percent_of_epc or fixed")
        add_param("Developer", "Fee Percentage", dev.get('developer_fee_pct', 0), "% of EPC")
        add_param("Developer", "Fee Fixed Amount", dev.get('developer_fee_fixed_usd', 0), "$")
        add_param("Developer", "Fee Timing", dev.get('developer_fee_timing', ''), "NTP, COD, over_time")
        row += 1

        # OPEX
        add_section("OPERATING EXPENDITURE")
        opex = self.inputs.get('opex', {})
        add_param("OpEx", "Fixed O&M", opex.get('fixed_om_usd_per_kw_year', 0), "$/kW-year")
        add_param("OpEx", "Variable O&M", opex.get('variable_om_usd_per_mwh', 0), "$/MWh")
        add_param("OpEx", "Insurance", opex.get('insurance_usd_per_kw_year', 0), "$/kW-year")
        add_param("OpEx", "Asset Management", opex.get('asset_mgmt_usd_per_kw_year', 0), "$/kW-year")
        add_param("OpEx", "Inverter Repairs", opex.get('inverter_repairs_major_usd_year', 0), "$/year")
        add_param("OpEx", "Other OpEx", opex.get('other_opex_usd_year', 0), "$/year")
        row += 1

        # LAND
        add_section("LAND")
        land = self.inputs.get('land', {})
        add_param("Land", "Mode", land.get('mode', ''), "lease or purchase")
        add_param("Land", "Lease Base Rate", land.get('lease_base_usd_per_acre_year', 0), "$/acre-year")
        add_param("Land", "Acres", land.get('acres', 0), "acres")
        add_param("Land", "Lease Escalator", land.get('lease_escalator_pct', 0), "% per year")
        add_param("Land", "Purchase Price", land.get('purchase_price_usd', 0), "$")
        row += 1

        # PROPERTY TAX
        add_section("PROPERTY TAX / PILOT")
        prop_tax = self.inputs.get('property_tax_pilot', {})
        add_param("Property Tax", "PILOT Enabled", prop_tax.get('pilot_enabled', False), "true/false")
        if prop_tax.get('pilot_enabled'):
            add_param("Property Tax", "PILOT Schedule", "See array below", "$/year by year")
        else:
            no_pilot = prop_tax.get('property_tax_if_no_pilot', {})
            add_param("Property Tax", "Assessed Value", no_pilot.get('assessed_value_usd', 0), "$")
            add_param("Property Tax", "Mill Rate", no_pilot.get('mill_rate_pct', 0), "%")
        row += 1

        # FINANCING
        add_section("FINANCING")
        fin = self.inputs.get('financing', {})
        add_param("Financing", "Use Debt", fin.get('use_debt', False), "true/false")
        add_param("Financing", "Sizing Method", fin.get('sizing_method', ''), "target_dscr or percent_of_cost")
        add_param("Financing", "Target Min DSCR", fin.get('target_min_dscr', 0), "ratio")
        add_param("Financing", "Tenor", fin.get('tenor_years', 0), "years")
        add_param("Financing", "Interest Rate", fin.get('interest_rate_pct', 0), "% per year")
        add_param("Financing", "Amortization", fin.get('amortization', ''), "sculpted or level")
        add_param("Financing", "Upfront Fees", fin.get('upfront_fees_pct_of_debt', 0), "% of debt")
        add_param("Financing", "DSRA", fin.get('dsra_months', 0), "months of debt service")
        add_param("Financing", "O&M Reserve", fin.get('om_reserve_months', 0), "months of O&M")
        row += 1

        # STATE PROGRAM
        add_section("STATE PROGRAM")
        prog = self.inputs.get('program', {})
        add_param("Program", "Active", prog.get('active', 'none'), "none, ny_vder, nj_csep, il_abp")

        if prog.get('active') == 'ny_vder':
            ws[f'A{row}'] = "(See NY-Sun/VDER details in JSON or Notes tab)"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
        elif prog.get('active') == 'nj_csep':
            ws[f'A{row}'] = "(See NJ CSEP details in JSON or Notes tab)"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
        elif prog.get('active') == 'il_abp':
            ws[f'A{row}'] = "(See IL ABP details in JSON or Notes tab)"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25

        # Apply borders to data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row_num in range(3, row):
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row_num}']
                cell.border = thin_border

                # Number formatting for value column
                if col == 'C' and row_num > 3:
                    if isinstance(cell.value, (int, float)) and cell.value != 0:
                        if cell.value > 1000:
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '0.00'

    def _create_energy_tab(self, wb: Workbook):
        """Create Energy_8760_or_Monthly tab."""
        ws = wb.create_sheet("Energy_8760_or_Monthly")

        # Write energy data
        self._write_dataframe_to_sheet(ws, self.cashflow[[
            'period', 'date', 'month_in_operation', 'year_in_operation',
            'ac_kwh', 'ac_mwh'
        ]], start_row=1)

    def _create_revenue_tab(self, wb: Workbook):
        """Create Revenue tab."""
        ws = wb.create_sheet("Revenue")

        # Identify revenue columns
        revenue_cols = ['period', 'date', 'month_in_operation', 'total_revenue']

        # Add specific revenue columns if they exist
        potential_cols = [
            'rec_revenue', 'capacity_revenue',
            'subscriber_revenue', 'ppa_revenue', 'merchant_revenue',
            'vder_energy', 'vder_environmental', 'vder_icap', 'vder_drv', 'vder_lsrv',
            'ny_community_credit', 'nj_bill_credit_revenue', 'nj_program_adders',
            'il_abp_rec_revenue', 'il_abp_brownfield_adder',
            'cdg_mgmt_fees', 'cdg_acquisition_costs', 'cdg_bad_debt',
            'ptc_amount'
        ]

        for col in potential_cols:
            if col in self.cashflow.columns:
                revenue_cols.append(col)

        self._write_dataframe_to_sheet(ws, self.cashflow[revenue_cols], start_row=1)

    def _create_capex_tab(self, wb: Workbook):
        """Create CapEx tab."""
        ws = wb.create_sheet("CapEx")

        # Summary
        ws['A1'] = "Capital Expenditure Summary"
        ws['A1'].font = Font(size=12, bold=True)

        ws['A3'] = "Total CapEx"
        ws['B3'] = self.metrics.get('total_capex', 0)
        ws['B3'].number_format = '$#,##0'

        # Write monthly schedule (subset of columns)
        capex_cols = ['period', 'date']
        if 'total_capex' in self.cashflow.columns:
            capex_cols.append('total_capex')

        self._write_dataframe_to_sheet(ws, self.cashflow[capex_cols], start_row=6)

    def _create_opex_tab(self, wb: Workbook):
        """Create OpEx tab."""
        ws = wb.create_sheet("OpEx")

        opex_cols = ['period', 'date', 'month_in_operation']

        potential_cols = [
            'fixed_om', 'variable_om', 'insurance', 'asset_management',
            'other_opex', 'land_lease', 'property_tax', 'inverter_replacement',
            'total_opex'
        ]

        for col in potential_cols:
            if col in self.cashflow.columns:
                opex_cols.append(col)

        self._write_dataframe_to_sheet(ws, self.cashflow[opex_cols], start_row=1)

    def _create_debt_tab(self, wb: Workbook):
        """Create Debt_Finance tab."""
        ws = wb.create_sheet("Debt_Finance")

        # Summary
        ws['A1'] = "Debt Financing Summary"
        ws['A1'].font = Font(size=12, bold=True)

        ws['A3'] = "Total Debt"
        ws['B3'] = self.metrics.get('total_debt', 0)
        ws['B3'].number_format = '$#,##0'

        ws['A4'] = "Min DSCR"
        ws['B4'] = self.metrics.get('min_dscr', 0)
        ws['B4'].number_format = '0.00'

        ws['A5'] = "Avg DSCR"
        ws['B5'] = self.metrics.get('avg_dscr', 0)
        ws['B5'].number_format = '0.00'

        # Write debt schedule
        debt_cols = ['period', 'date', 'debt_balance', 'interest_payment',
                    'principal_payment', 'total_debt_service']

        if 'dscr' in self.cashflow.columns:
            debt_cols.append('dscr')

        self._write_dataframe_to_sheet(ws, self.cashflow[debt_cols], start_row=8)

    def _create_tax_depreciation_tab(self, wb: Workbook):
        """Create Taxes_Depreciation tab."""
        ws = wb.create_sheet("Taxes_Depreciation")

        tax_cols = ['period', 'date', 'month_in_operation']

        potential_cols = [
            'depreciation_federal', 'depreciation_state',
            'taxable_income', 'federal_tax', 'state_tax', 'total_tax',
            'itc_credit', 'ptc_amount', 'upfront_grants',
            'nol_balance'
        ]

        for col in potential_cols:
            if col in self.cashflow.columns:
                tax_cols.append(col)

        self._write_dataframe_to_sheet(ws, self.cashflow[tax_cols], start_row=1)

    def _create_cashflow_waterfall_tab(self, wb: Workbook):
        """Create Cashflow_Waterfall tab."""
        ws = wb.create_sheet("Cashflow_Waterfall")

        waterfall_cols = [
            'period', 'date', 'month_in_operation', 'year_in_operation',
            'total_revenue', 'total_opex', 'ebitda', 'depreciation_federal',
            'ebit', 'interest_payment', 'taxable_income', 'total_tax',
            'itc_credit', 'upfront_grants', 'cfads', 'total_debt_service',
            'debt_drawdown', 'total_capex', 'fcfe',
            'equity_contribution', 'equity_distribution', 'terminal_value'
        ]

        # Filter to existing columns
        available_cols = [col for col in waterfall_cols if col in self.cashflow.columns]

        self._write_dataframe_to_sheet(ws, self.cashflow[available_cols], start_row=1)

    def _create_irr_npv_metrics_tab(self, wb: Workbook):
        """Create IRR_NPV_Metrics tab."""
        ws = wb.create_sheet("IRR_NPV_Metrics")

        ws['A1'] = "Financial Metrics Summary"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        # Returns
        ws[f'A{row}'] = "RETURNS"
        self._apply_section_style(ws[f'A{row}'])
        row += 1

        metrics_data = [
            ("Equity Pre-Tax IRR", self.metrics.get('pre_tax_irr', 0), '0.00%'),
            ("Equity Post-Tax IRR", self.metrics.get('post_tax_irr', 0), '0.00%'),
            ("Project IRR", self.metrics.get('project_irr', 0), '0.00%'),
            ("Equity NPV", self.metrics.get('equity_npv', 0), '$#,##0'),
            ("Project NPV", self.metrics.get('project_npv', 0), '$#,##0'),
            ("Payback (years)", self.metrics.get('payback_years', 0), '0.0'),
        ]

        for label, value, fmt in metrics_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = fmt
            row += 1

        row += 1

        # LCOE
        ws[f'A{row}'] = "LCOE"
        self._apply_section_style(ws[f'A{row}'])
        row += 1

        ws[f'A{row}'] = "Nominal LCOE ($/MWh)"
        ws[f'B{row}'] = self.metrics.get('nominal_lcoe', 0)
        ws[f'B{row}'].number_format = '$0.00'
        row += 1

        ws[f'A{row}'] = "Real LCOE ($/MWh)"
        ws[f'B{row}'] = self.metrics.get('real_lcoe', 0)
        ws[f'B{row}'].number_format = '$0.00'
        row += 1

        row += 1

        # Debt metrics
        ws[f'A{row}'] = "DEBT METRICS"
        self._apply_section_style(ws[f'A{row}'])
        row += 1

        debt_metrics = [
            ("Total Debt", self.metrics.get('total_debt', 0), '$#,##0'),
            ("Min DSCR", self.metrics.get('min_dscr', 0), '0.00'),
            ("Avg DSCR", self.metrics.get('avg_dscr', 0), '0.00'),
            ("Year 1 DSCR", self.metrics.get('year1_dscr', 0), '0.00'),
        ]

        for label, value, fmt in debt_metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value if not pd.isna(value) else "N/A"
            if not pd.isna(value):
                ws[f'B{row}'].number_format = fmt
            row += 1

        row += 1

        # Investment summary
        ws[f'A{row}'] = "INVESTMENT SUMMARY"
        self._apply_section_style(ws[f'A{row}'])
        row += 1

        inv_metrics = [
            ("Total CapEx", self.metrics.get('total_capex', 0), '$#,##0'),
            ("Total Equity Invested", self.metrics.get('total_equity_invested', 0), '$#,##0'),
            ("Lifetime Energy (MWh)", self.metrics.get('lifetime_energy_mwh', 0), '#,##0'),
            ("Lifetime Revenue", self.metrics.get('lifetime_revenue', 0), '$#,##0'),
        ]

        for label, value, fmt in inv_metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = fmt
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_sensitivities_tab(self, wb: Workbook):
        """Create Sensitivities tab."""
        ws = wb.create_sheet("Sensitivities")

        ws['A1'] = "Sensitivity Analysis"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = "Sensitivity analysis placeholder"
        ws['A4'] = "Run sensitivity scenarios to populate this tab"

        # Placeholder table
        headers = ['Parameter', 'Low', 'Base', 'High', 'IRR Impact']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col, value=header)
            self._apply_header_style(cell)

    def _create_scenarios_tab(self, wb: Workbook):
        """Create Scenarios tab."""
        ws = wb.create_sheet("Scenarios")

        ws['A1'] = "Scenario Analysis"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = "Scenarios:"
        scenarios = [
            "Base Case",
            "P90 Energy",
            "High CapEx (+20%)",
            "Low PPA Price (-10%)",
            "High Debt (70% LTV)"
        ]

        for i, scenario in enumerate(scenarios, start=4):
            ws[f'A{i}'] = scenario

    def _create_charts_tab(self, wb: Workbook):
        """Create Charts tab."""
        ws = wb.create_sheet("Charts")

        ws['A1'] = "Financial Charts"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = "Charts will be generated here:"
        ws['A4'] = "- Revenue stack (Year 1)"
        ws['A5'] = "- Cashflow waterfall"
        ws['A6'] = "- Debt balance over time"
        ws['A7'] = "- DSCR over time"
        ws['A8'] = "- IRR sensitivity tornado"

    def _create_audit_trace_tab(self, wb: Workbook):
        """Create Audit_Trace tab."""
        ws = wb.create_sheet("Audit_Trace")

        ws['A1'] = "Audit Trail and Assumptions Log"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        # Defaults used
        ws[f'A{row}'] = "DEFAULTS APPLIED"
        self._apply_section_style(ws[f'A{row}'])
        row += 1

        if self.defaults_used:
            for default in self.defaults_used:
                ws[f'A{row}'] = default
                row += 1
        else:
            ws[f'A{row}'] = "No defaults applied - all inputs provided"
            row += 1

        row += 1

        # Warnings
        ws[f'A{row}'] = "WARNINGS"
        self._apply_section_style(ws[f'A{row}'])
        row += 1

        if self.warnings:
            for warning in self.warnings:
                ws[f'A{row}'] = warning
                row += 1
        else:
            ws[f'A{row}'] = "No warnings"
            row += 1

        ws.column_dimensions['A'].width = 80

    def _create_notes_tab(self, wb: Workbook):
        """Create Notes tab."""
        ws = wb.create_sheet("Notes")

        ws['A1'] = "Model Notes and Documentation"
        ws['A1'].font = Font(size=14, bold=True)

        notes = [
            "",
            "UNITS AND CONVENTIONS:",
            "- Currency: USD",
            "- Energy: kWh, MWh as labeled",
            "- Capacity: kW (AC unless specified DC)",
            "- Time: Monthly periods from NTP through end of operating life",
            "- Rates: Annual unless specified monthly",
            "",
            "ASSUMPTIONS:",
            "- All inputs are user-provided and documented in Inputs_Data tab",
            "- State program values (NY-Sun, NJ CSEP, IL ABP) are user inputs and may change by utility/zone/vintage",
            "- This model does NOT fetch live tariff rates or REC prices",
            "- Tax credits and depreciation follow federal rules as of model creation date",
            "- NOL carryforward is unlimited in time (conservative assumption)",
            "",
            "METHODOLOGY:",
            "- Energy: Monthly production with annual degradation",
            "- Revenue: Mode-specific (Community Solar vs PPA) with program overlays",
            "- Tax: ITC/PTC mutually exclusive, MACRS depreciation with basis reduction",
            "- Debt: Sized to target DSCR or % of cost, sculpted or level amortization",
            "- Metrics: XIRR/XNPV on monthly cashflows",
            "",
            "LIMITATIONS:",
            "- No tax equity partnership modeling (can be added)",
            "- Simplified subscriber churn model for Community Solar",
            "- Merchant pricing uses last PPA price as default",
            "- Property tax/PILOT uses simple escalation",
            "",
            "For questions or to report issues, see README.md"
        ]

        for i, note in enumerate(notes, start=3):
            ws[f'A{i}'] = note

        ws.column_dimensions['A'].width = 100

    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, start_row: int = 1):
        """Helper to write DataFrame to sheet with formatting."""
        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            self._apply_header_style(cell)

        # Write data
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply number formatting
                if col_idx > 1:  # Skip period/date columns
                    col_name = df.columns[col_idx - 1]

                    if 'revenue' in col_name.lower() or 'cost' in col_name.lower() or \
                       'capex' in col_name.lower() or 'opex' in col_name.lower() or \
                       'tax' in col_name.lower() or 'debt' in col_name.lower() or \
                       'equity' in col_name.lower() or 'ebitda' in col_name.lower() or \
                       'npv' in col_name.lower() or 'payment' in col_name.lower():
                        cell.number_format = '#,##0'
                    elif 'irr' in col_name.lower() or 'pct' in col_name.lower():
                        cell.number_format = '0.00%'
                    elif 'dscr' in col_name.lower():
                        cell.number_format = '0.00'
                    elif 'kwh' in col_name.lower() or 'mwh' in col_name.lower():
                        cell.number_format = '#,##0.0'

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _apply_header_style(self, cell):
        """Apply header style to cell."""
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def _apply_section_style(self, cell):
        """Apply section header style to cell."""
        cell.fill = self.section_fill
        cell.font = self.section_font
